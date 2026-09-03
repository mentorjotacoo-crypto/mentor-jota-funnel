# -*- coding: utf-8 -*-
"""
Extrae el detalle de citas de los Consolidados Closer (Drive local) y genera:
  - closers.json  (local, gitignored — para inspección)
  - closers.enc   (AES-256-GCM, SÍ se commitea — build.js lo fusiona al payload)

Los 5 archivos tienen 3 layouts distintos y algunas hojas se repiten entre
libros (Jose Luis Aldea y Jota están en dos), así que el mapeo es POR NOMBRE
de columna y la deduplicación por (fecha, cliente, correo, hoja).

Uso:  python sync_closers.py "clave"
Sale con código 9 si no hay cambios respecto al closers.json previo.
"""
import sys
import os
import json
import gzip
import base64
import glob
import warnings
from datetime import datetime

warnings.filterwarnings('ignore')          # openpyxl: data validation extension
import openpyxl
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes

CARPETA = r'H:\Mi unidad\Empresas Cynthia y Jose Julian\Mentor Jota\Técnico\Closer'
REPO = os.path.dirname(os.path.abspath(__file__))
ITERATIONS = 150000

SALTAR_HOJAS = {'INPUTS', 'Claude Log', 'Hoja1', 'Hoja2'}

# Un cierre es una venta real. 'Compromiso'/'Pendiente' son promesas, NO cuentan.
CIERRE = {'si', 'abona', 'paga cuotas'}

# Alias de columna -> campo interno (se compara con startswith en minúsculas)
CAMPOS = {
    'fecha':  ['fecha'],
    'closer': ['colaborador'],
    'tipo':   ['tipo agendamien'],
    'origen': ['agendamiento'],
    'nombre': ['nombre cliente'],
    'correo': ['correo'],
    'pais':   ['pais', 'país'],
    'asist':  ['asistencia'],
    'estado': ['estado cita', 'estado de la cita'],
    'compra': ['compra'],
    'motivo': ['motivo cierre'],
    'prog':   ['programa'],
    'venta':  ['valor venta'],
    'pagado': ['valor pagado'],
}

# El mismo país llega escrito de varias formas entre libros
PAIS_NORM = {
    'mexico': 'México', 'méxico': 'México',
    'estados unidos': 'Estados Unidos', 'usa': 'Estados Unidos', 'eeuu': 'Estados Unidos',
    'canada': 'Canadá', 'canadá': 'Canadá',
    'peru': 'Perú', 'perú': 'Perú',
    'espana': 'España', 'españa': 'España',
    'republica dominicana': 'República Dominicana',
    'república dominicana': 'República Dominicana',
}

# Nombres que aparecen distinto en la columna Colaborador vs el nombre de hoja
CLOSER_NORM = {
    'christian': 'Christian Correa',
    'jose julian valencia': 'Jota',
    'jose julian valen': 'Jota',
}


def num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace('$', '').replace(',', '').strip() or 0)
    except ValueError:
        return 0.0


def norm_pais(v):
    s = str(v or '').strip()
    if not s:
        return ''
    return PAIS_NORM.get(s.lower(), s)


def norm_closer(v, hoja):
    s = str(v or '').strip()
    if not s:
        return hoja
    return CLOSER_NORM.get(s.lower(), s)


def leer():
    filas = []
    vistos = set()
    archivos = [f for f in sorted(glob.glob(os.path.join(CARPETA, '*.xlsx')))
                if not os.path.basename(f).startswith('~$')]
    if not archivos:
        print('[SKIP] no se encontraron consolidados (¿Drive sin sincronizar?)')
        sys.exit(0)

    for f in archivos:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if sn in SALTAR_HOJAS:
                continue
            ws = wb[sn]
            try:
                cab = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                continue
            hdr = [str(c).strip().lower() if c else '' for c in cab]
            ix = {}
            for i, h in enumerate(hdr):
                for campo, pats in CAMPOS.items():
                    if campo not in ix and any(h.startswith(p) for p in pats):
                        ix[campo] = i
            if 'fecha' not in ix:
                continue

            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r:
                    continue
                fx = r[ix['fecha']] if ix['fecha'] < len(r) else None
                if not isinstance(fx, datetime):
                    continue
                g = lambda k: (r[ix[k]] if k in ix and ix[k] < len(r) else None)
                nombre = str(g('nombre') or '').strip()
                correo = str(g('correo') or '').strip().lower()
                clave = (fx.date().isoformat(), nombre.lower(), correo, sn)
                if clave in vistos:
                    continue
                vistos.add(clave)

                asistio = str(g('asist') or '').strip().lower().startswith('s')
                compra_raw = str(g('compra') or '').strip().lower()
                cerro = compra_raw in CIERRE
                filas.append({
                    'fecha': fx.date().isoformat(),
                    'closer': norm_closer(g('closer'), sn),
                    'pais': norm_pais(g('pais')),
                    'tipo': str(g('tipo') or '').strip(),
                    'origen': str(g('origen') or '').strip(),
                    'asistio': asistio,
                    'cerro': cerro,
                    'motivo': str(g('motivo') or '').strip(),
                    'programa': str(g('prog') or '').strip(),
                    'venta': num(g('venta')) if cerro else 0.0,
                    'pagado': num(g('pagado')) if cerro else 0.0,
                })
        wb.close()
    return filas


def compactar(filas):
    """Formato columnar con diccionarios de índices: pesa ~6x menos que
    la lista de objetos y comprime mucho mejor."""
    closers, paises, tipos, motivos, progs = [], [], [], [], []

    def idx(lista, v):
        if v not in lista:
            lista.append(v)
        return lista.index(v)

    filas = sorted(filas, key=lambda x: x['fecha'])
    r = []
    for f in filas:
        r.append([
            f['fecha'],
            idx(closers, f['closer']),
            idx(paises, f['pais']),
            1 if f['asistio'] else 0,
            1 if f['cerro'] else 0,
            round(f['venta']),
            round(f['pagado']),
            idx(tipos, f['tipo']),
            idx(motivos, f['motivo']),
            idx(progs, f['programa']),
        ])
    return {
        'cols': ['fecha', 'closer', 'pais', 'asistio', 'cerro', 'venta',
                 'pagado', 'tipo', 'motivo', 'programa'],
        'closers': closers, 'paises': paises, 'tipos': tipos,
        'motivos': motivos, 'programas': progs,
        'r': r,
        'actualizado': datetime.now().isoformat(timespec='seconds'),
    }


def main():
    if len(sys.argv) < 2 or len(sys.argv[1]) < 8:
        print('Uso: python sync_closers.py "clave"  (min 8 chars)')
        sys.exit(1)
    password = sys.argv[1]

    print('Leyendo consolidados...')
    filas = leer()
    if not filas:
        print('[SKIP] no se extrajo ninguna cita')
        sys.exit(0)

    data = compactar(filas)
    tot = len(data['r'])
    asis = sum(f[3] for f in data['r'])
    cie = sum(f[4] for f in data['r'])
    print(f'  Citas: {tot} | asistidas: {asis} ({asis/tot*100:.1f}%) | '
          f'cierres: {cie} ({cie/asis*100:.1f}%% s/asistidas)'.replace('%%','%'))
    print(f'  Closers: {len(data["closers"])} | paises: {len(data["paises"])}')
    print(f'  Rango: {data["r"][0][0]} a {data["r"][-1][0]}')

    # Idempotencia: si el detalle no cambió, no republicar (el .enc siempre
    # difiere byte a byte por el salt aleatorio).
    jpath = os.path.join(REPO, 'closers.json')
    if os.path.exists(jpath):
        try:
            with open(jpath, encoding='utf-8') as f:
                prev = json.load(f)
            if prev.get('r') == data['r']:
                print('[SKIP] sin cambios respecto al último sync')
                sys.exit(9)
        except (OSError, ValueError):
            pass

    with open(jpath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

    plaintext = gzip.compress(json.dumps(data, ensure_ascii=False).encode('utf-8'), 9)
    salt = os.urandom(16)
    iv = os.urandom(12)
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=ITERATIONS)
    key = kdf.derive(password.encode('utf-8'))
    ct_tag = AESGCM(key).encrypt(iv, plaintext, None)
    ct, tag = ct_tag[:-16], ct_tag[-16:]
    payload = base64.b64encode(salt + iv + tag + ct).decode('ascii')

    with open(os.path.join(REPO, 'closers.enc'), 'w', encoding='utf-8') as f:
        f.write(payload)
    print(f'[OK] closers.enc escrito ({len(payload)/1024:.1f} KB · '
          f'{len(plaintext)/1024:.1f} KB comprimidos)')


if __name__ == '__main__':
    main()
