# -*- coding: utf-8 -*-
"""
Extrae los datos diarios de ascensos (high ticket) de la Hoja de seguimiento
Mentor Jota Consolidado.xlsx (Drive local) y genera:
  - ascensos.json  (local, gitignored — para inspección)
  - ascensos.enc   (encriptado AES-256-GCM, SÍ se commitea — build.js lo
                    desencripta y fusiona al payload del dashboard)

Uso:  python sync_ascensos.py "clave"
"""
import sys
import json
import base64
import gzip
import os
from datetime import datetime, date

import openpyxl
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes

XLSX = r'H:\Mi unidad\Empresas Cynthia y Jose Julian\Mentor Jota\Técnico\Hoja de seguimiento Mentor Jota Consolidado.xlsx'
REPO = os.path.dirname(os.path.abspath(__file__))

MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio',
         'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

# Columnas (1-based) en las pestañas mensuales — layout idéntico 2025/2026
# Tres modalidades con precio/pago separados:
#   Mentoría (principal): ventas/contratado/cobrado
#   STM Curso y STM Lite: sus propias columnas de venta, precio y pago
COLS = {
    'spend': 3,
    'impressions': 4,
    'clicks': 7,
    'leads': 10,
    'agendamientos': 13,
    'llamadas': 17,
    'ventas': 20,               # Mentoría
    'contratado': 26,           # Precio venta Mentoría
    'cobrado': 27,              # Pago Mentoría
    'ventas_stm_curso': 22,
    'contratado_stm_curso': 29,
    'cobrado_stm_curso': 31,
    'ventas_stm_lite': 23,
    'contratado_stm_lite': 30,
    'cobrado_stm_lite': 32,
}

# Mismo formato que build.js: PBKDF2-SHA256 150k iters, salt(16)|iv(12)|tag(16)|ct
ITERATIONS = 150000


def num(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace('$', '').replace(',', '').strip() or 0)
    except ValueError:
        return 0


def main():
    if len(sys.argv) < 2 or len(sys.argv[1]) < 8:
        print('Uso: python sync_ascensos.py "clave"  (min 8 chars)')
        sys.exit(1)
    password = sys.argv[1]

    print(f'Leyendo {XLSX}...')
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    hoy = date.today()
    days = {}
    hojas_leidas = []

    for anio in (2025, 2026):
        for mi, mes in enumerate(MESES):
            nombre = f'{mes} {anio}'
            if nombre not in wb.sheetnames:
                continue
            ws = wb[nombre]
            filas = 0
            for row in ws.iter_rows(min_row=2, max_col=max(COLS.values())):
                dia_cell = row[0].value
                if not isinstance(dia_cell, datetime):
                    continue
                d = dia_cell.date()
                if d.year != anio or d.month != mi + 1:
                    continue  # basura fuera del mes
                if d > hoy:
                    continue  # días futuros del mes en curso
                key = d.isoformat()
                if key not in days:
                    days[key] = {k: 0 for k in COLS}
                # Sumar (soporta varias fuentes por día: Facebook, Google...)
                for k, c in COLS.items():
                    days[key][k] += num(row[c - 1].value)
                filas += 1
            hojas_leidas.append(f'{nombre} ({filas} filas)')

    # Redondear para no arrastrar floats largos
    for k, v in days.items():
        for m in v:
            v[m] = round(v[m], 2)

    ascensos = {
        'updatedAt': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'source': 'Hoja de seguimiento Mentor Jota Consolidado.xlsx',
        'days': days,
    }

    print(f'Hojas leídas: {len(hojas_leidas)}')
    for h in hojas_leidas:
        print(f'  - {h}')
    print(f'Días totales: {len(days)}')

    # Guardar JSON local (gitignored)
    json_path = os.path.join(REPO, 'ascensos.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(ascensos, f, ensure_ascii=False, indent=2)
    print(f'[OK] {json_path}')

    # Encriptar → ascensos.enc (mismo formato que el payload del dashboard)
    # Comprimir antes de cifrar: build.js detecta el magic gzip y descomprime.
    plaintext = gzip.compress(json.dumps(ascensos, ensure_ascii=False).encode('utf-8'), 9)
    salt = os.urandom(16)
    iv = os.urandom(12)
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=ITERATIONS)
    key = kdf.derive(password.encode('utf-8'))
    ct_tag = AESGCM(key).encrypt(iv, plaintext, None)  # ct + tag(16) al final
    ct, tag = ct_tag[:-16], ct_tag[-16:]
    payload = base64.b64encode(salt + iv + tag + ct).decode('ascii')

    enc_path = os.path.join(REPO, 'ascensos.enc')
    with open(enc_path, 'w', encoding='ascii') as f:
        f.write(payload)
    print(f'[OK] {enc_path} ({len(payload)} chars)')
    print('Siguiente: node build.js "clave" && git add ascensos.enc index.html && commit+push')


if __name__ == '__main__':
    main()
